import pandas as pd
from openpyxl import load_workbook

df = pd.read_csv('practice\ltx.csv')
df_group = df.groupby(['PurposeOfTheAssignment', 'Koatyy','Price','ValueNGO'], as_index=False)
a = df['PurposeOfTheAssignment'].value_counts().reset_index()
b = df.groupby('PurposeOfTheAssignment')['Koatyy'].sum().sort_values(ascending=False).reset_index()
c = df.groupby('PurposeOfTheAssignment')['Price'].sum().sort_values(ascending=False).reset_index()
d = df.groupby('PurposeOfTheAssignment')['ValueNGO'].sum().sort_values(ascending=False).reset_index()


a1 = a['index']                   
a1_1 = a['PurposeOfTheAssignment']
b1 = b['Koatyy']                  
c1 = c['Price']                   
d1 = d['ValueNGO']                
length = a.shape[0]               


c2 = c1.div(b1, fill_value=0)#
d2 = c1.div(d1, fill_value=0)#

print('-----------INDEX---------------')                       
print(a1)                                                      
print('-----------PurposeOfTheAssignment---------------')      
print(a1_1)                                                    
print('------------Koatyy--------------')                      
print(b1)                                                      
print('-------------Price-------------')                       
print(c2)                                                      
print('--------------ValueNGO------------')                    
print(d2)                                                      
print('--------------------------')                            


data = {'Цільове призначення': a1 , 'Кількість угод': a1_1, 'Площа, га': b1, 'Ціна (вартість), грн./га': c2, 'Значення НГО, грн./га': d2}
df = pd.DataFrame(data)

workbook = load_workbook(filename='practice\opus.xlsx')
worksheet = workbook['123']

last_row = worksheet.max_row
while worksheet.cell(row=last_row, column=2).value is None and last_row > 1:
    last_row -= 1

arc = last_row+1
print(last_row)


with pd.ExcelWriter('opus.xlsx') as writer:
    pd.DataFrame([length]).to_excel(writer, sheet_name='123',startrow=arc,startcol=1, header=False,index =False)
    df.to_excel(writer, sheet_name='123', index=True,startrow=8,startcol=0)
    
print(worksheet.max_row)
worksheet.cell(row=arc-1, column=1, value='Всього')
worksheet.cell(row=9, column=1, value='№ п/п')
workbook.save('opus.xlsx')
